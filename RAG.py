import logging
import mimetypes
import os
import re
from concurrent.futures import ThreadPoolExecutor

import cv2
import fitz
import numpy as np
import openpyxl
import pytesseract
import requests
from PIL import Image
from bs4 import BeautifulSoup
from colorama import Fore, Style, init
from easyocr import easyocr
from langchain.chains import RetrievalQA
from langchain.prompts import PromptTemplate
from langchain_community.vectorstores import Chroma
from langchain_core.embeddings import Embeddings
from langchain_ollama import ChatOllama
from langchain_text_splitters import RecursiveCharacterTextSplitter
from sentence_transformers import SentenceTransformer
from yolov5 import YOLOv5

from tqdm import tqdm
tqdm.disable = True  # 全局禁用进度条

# Initialize colorama for colored console output
init(autoreset=True)

# Set environment variables
os.environ['USER_AGENT'] = 'FinTechHelper/1.0 (xzco.work@gmail.com)'
os.environ["TOKENIZERS_PARALLELISM"] = "false"
os.environ["LANGCHAIN_TRACING_V2"] = "true"
os.environ["LANGCHAIN_PROJECT"] = "RAG"
os.environ["LANGCHAIN_API_KEY"] = "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"

COLOR_GREEN = Fore.GREEN
COLOR_RED = Fore.RED
COLOR_CYAN = Fore.CYAN
COLOR_YELLOW = Fore.YELLOW
COLOR_MAGENTA = Fore.MAGENTA
COLOR_WHITE = Fore.WHITE
RESET = Style.RESET_ALL


# 文件加载与预处理
def load_and_process_files(files_path_lists):
    texts = []
    for file_path in files_path_lists:
        raw_text = load_text_from_file(file_path)
        cleaned_text = clean_text(raw_text)
        texts.extend(split_text(cleaned_text))
    return texts


# 文件路径获取
def attain_paths_of_all_files(directory_path):
    supported_extensions = (".txt", ".pdf", ".xlsx", ".csv", ".jpg", ".png")
    files_paths = []
    for root, dirs, filenames in os.walk(directory_path):
        for filename in filenames:
            if filename.endswith(supported_extensions):
                files_paths.append(os.path.join(root, filename))
    return files_paths


# 文件加载
def load_text_from_file(file_path):
    mime_type, _ = mimetypes.guess_type(file_path)
    if mime_type == 'text/plain':
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    elif mime_type == 'application/pdf':
        return extract_text_from_pdf(file_path)
    elif mime_type in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
        return extract_text_from_excel(file_path)
    elif mime_type and mime_type.startswith('image'):
        return extract_text_from_image(file_path)
    elif mime_type is None and file_path.startswith("http"):
        return extract_text_from_webpage(file_path)
    else:
        raise ValueError(f"Unsupported file type for source: {file_path}")


def extract_text_from_pdf(file_path):
    extracted_text = []
    with fitz.open(file_path) as pdf:
        for page in pdf:
            text = page.get_text("text")
            if text.strip():  # 如果提取到文本，直接使用
                extracted_text.append(text)
            else:  # 如果没有文本，使用 OCR
                pix = page.get_pixmap()
                # 将宽度和高度以元组形式传递
                image = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
                text = pytesseract.image_to_string(image, lang='eng')
                extracted_text.append(text)
    return "\n".join(extracted_text)


def extract_text_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    all_text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_text = f"Sheet: {sheet_name}\n"
        for row in sheet.iter_rows(values_only=True):
            row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
            sheet_text += row_text + "\n"
        all_text.append(sheet_text)
    return "\n".join(all_text)


def extract_text_from_image(file_path):
    yolo_model = YOLOv5('yolov5s.pt')
    # 1. 读取图片
    image = cv2.imread(file_path)

    # 2. 使用 YOLO 检测文本区域
    results = yolo_model.predict(image)  # YOLOv5返回的结果包括bounding boxes和类别
    text_boxes = results.xywh[0][:, :4].cpu().numpy()  # 获取检测到的文本区域（假设是第一个类）

    # 3. 从检测到的区域中提取文本
    reader = easyocr.Reader(['ch_sim', 'en'], gpu=True)
    extracted_text = ""
    for box in text_boxes:
        # 提取每个文本区域
        x1, y1, x2, y2 = map(int, box)  # 得到左上角和右下角坐标
        roi = image[y1:y2, x1:x2]  # 裁剪文本区域

        # 使用 OCR 对该区域进行文本识别
        result = reader.readtext(roi)
        for _, text, _ in result:
            extracted_text += text + "\n"

    return extracted_text


def extract_text_from_webpage(url):
    response = requests.get(url)
    response.encoding = 'utf-8'
    soup = BeautifulSoup(response.text, 'html.parser')
    return soup.get_text(separator="\n")


# 文本分割
def split_text(text, chunk_size=1000, chunk_overlap=150):
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        length_function=len
    )
    return text_splitter.create_documents([text])


# 文本清理
def clean_text(text, keep_punctuation=False):
    text = re.sub(r'</?[a-zA-Z][^>]*>', '', text)
    text = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', text)
    if not keep_punctuation:
        text = re.sub(r'[^\w\s]', '', text)
    else:
        text = re.sub(r'[^\w\s.,!?]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


# Embedding类
class SentenceTransformersEmbedding(Embeddings):
    def __init__(self, model_name: str):
        self.model = SentenceTransformer(model_name)

    def embed_documents(self, texts):
        embeddings = self.model.encode(texts, convert_to_tensor=False, show_progress_bar=False).tolist()
        if not embeddings:
            raise ValueError("No embeddings generated for the documents.")
        return embeddings

    def embed_query(self, text):
        embedding = self.model.encode([text], convert_to_tensor=False, show_progress_bar=False).tolist()[0]
        if not embedding:
            raise ValueError("No embedding generated for the query.")
        return embedding


# 初始化向量存储
def initialize_vector_store(texts, embedding_model, persist_directory):
    if os.path.exists(persist_directory):
        logging.info("Loading existing vector store...")
    else:
        logging.info("Creating new vector store...")
    return Chroma.from_documents(texts, embedding=embedding_model, persist_directory=persist_directory)


def calculate_similarity(query_embedding, doc_embedding):
    if doc_embedding is None:
        return 0.0  # 如果嵌入为空，直接返回最低分
    return np.dot(query_embedding, doc_embedding) / (np.linalg.norm(query_embedding) * np.linalg.norm(doc_embedding))


def find_best_documents(query_embedding, docs, top_k=3):
    scored_docs = []
    for doc in docs:
        doc_embedding = doc.metadata.get('embedding')
        if doc_embedding is not None:
            score = calculate_similarity(query_embedding, doc_embedding)
            scored_docs.append((score, doc))

    # 按分数排序并返回前 top_k 个文档
    scored_docs.sort(reverse=True, key=lambda x: x[0])
    best_docs = [doc for _, doc in scored_docs[:top_k]]
    return scored_docs[:top_k]  # 返回 (score, doc) 的元组列表


def print_answer_and_sources(agent_name, query, answer, source_documents, query_embedding):
    """打印问题、答案以及来源文档的相似度评分"""
    print(f"\n\033[94mAgent [{agent_name}] Processing:\033[0m")
    print(f"\n\033[96mQuestion:\033[0m\n{query}")
    print(f"\n\033[93mAnswer:\033[0m\n{answer}")
    print(f"\n\033[95mSource Documents with Similarity Scores:\033[0m")

    # 表头和分割线
    print(f"{'Document Excerpt':<60} {'Similarity Score':>10}")
    print("-" * 80)

    for doc in source_documents:
        doc_content = doc.page_content[:50] + "..." if len(doc.page_content) > 50 else doc.page_content
        doc_embedding = doc.metadata.get('embedding', None)
        score = calculate_similarity(query_embedding, doc_embedding)

        score_display = f"{score:.4f}" if score is not None else "N/A"
        print(f"{doc_content:<60} {score_display:>10}")

    print("-" * 80)


def process_query(qa_chain, embedding_model, retriever, query, similarity_threshold=0.5, top_k=3):
    """处理查询并打印问题、答案和来源文档信息"""
    # Step 1: 使用 QA Chain 获取初始答案和来源文档
    response = qa_chain.invoke({"query": query})
    source_documents = response["source_documents"]

    # Step 2: 生成查询的嵌入
    query_embedding = embedding_model.generate_embedding(query)

    # Step 3: 为文档生成嵌入（如果未生成）
    for doc in source_documents:
        if 'embedding' not in doc.metadata:
            doc.metadata['embedding'] = embedding_model.generate_embedding(doc.page_content)

    # Step 4: 找到最相似的文档
    best_docs_with_scores = find_best_documents(query_embedding, source_documents, top_k=top_k)
    best_docs = [doc for _, doc in best_docs_with_scores]

    # 打印问题、答案和来源文档
    print_answer_and_sources(query, response['result'], best_docs, query_embedding)

    # Step 5: 自我批评机制验证答案
    final_answer = self_critic_retrieval(
        query=query,
        embedding_model=embedding_model,
        generated_answer=response['result'],
        query_embedding=query_embedding,
        best_docs_with_scores=best_docs_with_scores,
        retriever=retriever,
        similarity_threshold=similarity_threshold,
    )

    print(f"\n\033[92mFinal Answer (after self-critic validation):\033[0m\n{final_answer}")
    return final_answer


def self_critic_retrieval(query, embedding_model, generated_answer, query_embedding, best_docs_with_scores, retriever,
                          similarity_threshold=0.5):
    # 获取最高分文档及其分数
    best_score, best_doc = best_docs_with_scores[0] if best_docs_with_scores else (0, None)

    # Step 1: 如果最相似文档的得分低于阈值，重新检索
    if best_score < similarity_threshold:
        print(f"Low similarity score ({best_score}). Retrieving more documents...")
        new_response = retriever({"query": query, "k": 5})
        new_source_documents = new_response["source_documents"]

        # 为新文档生成嵌入
        for doc in new_source_documents:
            if 'embedding' not in doc.metadata:
                doc.metadata['embedding'] = embedding_model.embed_query(doc.page_content)

        # 再次寻找最佳文档
        best_docs_with_scores = find_best_documents(query_embedding, new_source_documents)
        best_docs = [doc for _, doc in best_docs_with_scores]

        # 使用新的文档重新生成答案
        return new_response["result"]

    # Step 2: 如果得分较高，直接返回原答案
    return generated_answer


# Agents
class EmbeddingAgent:
    def __init__(self, embedding_model):
        self.embedding_model = embedding_model

    def generate_embedding(self, text):
        return self.embedding_model.embed_query(text)


class RetrievalAgent:
    def __init__(self, retriever):
        self.retriever = retriever

    def retrieve_documents(self, query, k=3):
        # 使用 retriever 的 invoke 方法替代 get_relevant_documents
        return self.retriever.invoke({"query": query, "k": k})


class QAAgent:
    def __init__(self, qa_chain):
        self.qa_chain = qa_chain

    def generate_answer(self, query, query_embedding, source_documents, agent_name="QAAgent"):
        # Step 1: 生成答案
        response = self.qa_chain.invoke({"query": query})
        answer = response["result"]

        # Step 2: 打印答案和来源文档
        print_answer_and_sources(agent_name, query, answer, source_documents, query_embedding)
        return answer




class CriticAgent:
    def __init__(self, similarity_threshold=0.5, retriever=None, semantic_retriever=None):
        self.similarity_threshold = similarity_threshold
        self.retriever = retriever  # 基础检索器（TF-IDF、BM25等）
        self.semantic_retriever = semantic_retriever  # 语义检索器（如 Sentence-BERT）

    def find_best_documents(self, query_embedding, docs, top_k=3):
        """
        在静态文档集合中进行相似度排序。
        """
        scored_docs = []
        for doc in docs:
            doc_embedding = doc.metadata.get('embedding')
            if doc_embedding is not None:
                # 使用余弦相似度计算
                score = self.calculate_similarity(query_embedding, doc_embedding)
                scored_docs.append((score, doc))

        # 按分数排序并返回前 top_k 个文档
        scored_docs.sort(reverse=True, key=lambda x: x[0])
        return scored_docs[:top_k]

    def validate_answer(self, query, query_embedding, best_docs_with_scores, qa_agent, agent_name="CriticAgent"):
        """
        检查初始文档集的质量，必要时动态扩展检索范围。
        """
        best_score, _ = best_docs_with_scores[0] if best_docs_with_scores else (0, None)

        # 如果相似度低于阈值，触发新的检索逻辑
        if best_score < self.similarity_threshold:
            print(f"\n\033[93m[{agent_name}] Low similarity score ({best_score:.4f}). Retrieving more documents...\033[0m")

            # 动态扩展：语义检索 + 多上下文
            alternative_contexts = ["economic analysis", "market trends"]
            additional_docs = []
            for context in alternative_contexts:
                extra_docs = self.semantic_retriever.retrieve_documents(query, context=context, k=5)
                additional_docs.extend(extra_docs)

            # 对扩展文档重新排序
            reranked_docs = self.find_best_documents(query_embedding, additional_docs, top_k=5)
            answer = qa_agent.generate_answer(query, query_embedding, [doc for _, doc in reranked_docs], agent_name)
            return answer

        print(f"\n\033[92m[{agent_name}] High similarity score ({best_score:.4f}). Keeping original answer.\033[0m")
        return None

    def calculate_similarity(self, query_embedding, doc_embedding):
        """
        使用余弦相似度计算相似性分数。
        """
        return np.dot(query_embedding, doc_embedding) / (
            np.linalg.norm(query_embedding) * np.linalg.norm(doc_embedding)
        )



class MultiAgentRAG:
    def __init__(self, embedding_agent, retrieval_agent, qa_agent, critic_agent):
        self.embedding_agent = embedding_agent
        self.retrieval_agent = retrieval_agent
        self.qa_agent = qa_agent
        self.critic_agent = critic_agent

    def process_query(self, query, top_k=3):
        print(f"\nProcessing query: {query}")

        # Step 1: Generate query embedding
        query_embedding = self.embedding_agent.generate_embedding(query)

        # Step 2: Retrieve documents
        source_documents = self.retrieval_agent.retrieve_documents(query, k=top_k)

        # Step 3: Generate embeddings for documents (if not already generated)
        for doc in source_documents:
            if 'embedding' not in doc.metadata:
                doc.metadata['embedding'] = self.embedding_agent.generate_embedding(doc.page_content)

        # Step 4: Find best documents
        best_docs_with_scores = find_best_documents(query_embedding, source_documents, top_k=top_k)
        best_docs = [doc for _, doc in best_docs_with_scores]

        # Step 5: Generate initial answer
        initial_answer = self.qa_agent.generate_answer(query, query_embedding, best_docs)

        # Step 6: Validate answer using critic agent
        final_answer = self.critic_agent.validate_answer(
            query=query,
            query_embedding=query_embedding,
            best_docs_with_scores=best_docs_with_scores,
            qa_agent=self.qa_agent
        ) or initial_answer

        # Output final answer
        print(f"\n\033[92mFinal Answer:\033[0m\n{final_answer}")
        return final_answer


# Main Function
def main():
    llm = ChatOllama(model="llama3.1:latest")
    files_path_lists = attain_paths_of_all_files(
        "/Users/xzc/Library/Mobile Documents/com~apple~CloudDocs/CUHK/CMSC5720I-Project/LLM/data")
    texts = load_and_process_files(files_path_lists)

    # Initialize agents
    embedding_model = SentenceTransformersEmbedding("sentence-transformers/all-MiniLM-L12-v2")
    vector_store = initialize_vector_store(texts, embedding_model, persist_directory="./vector_store")
    retriever = vector_store.as_retriever(search_type="similarity", search_kwargs={"k": 3})

    prompt_template = PromptTemplate(
        input_variables=["context", "question"],
        template="""
            You are a helpful assistant that answers questions based on provided context.

            The following documents were retrieved to help answer the question:
            {context}

            Question: {question}

            Answer: If the answer can be found in the context, provide it. If the answer cannot be determined from the context, respond with "Cannot find the context."
        """
    )

    qa_chain = RetrievalQA.from_chain_type(
        llm=llm,
        chain_type="stuff",
        retriever=retriever,
        return_source_documents=True,
        chain_type_kwargs={"prompt": prompt_template}
    )

    embedding_agent = EmbeddingAgent(embedding_model)
    retrieval_agent = RetrievalAgent(retriever)
    qa_agent = QAAgent(qa_chain)
    critic_agent = CriticAgent(similarity_threshold=0.5)

    # Initialize Multi-Agent RAG system
    multi_agent_rag = MultiAgentRAG(embedding_agent, retrieval_agent, qa_agent, critic_agent)

    # Process queries
    queries = ["Please analyze Apple Inc.'s Operating expenses in detail"]

    with ThreadPoolExecutor(max_workers=10) as executor:
        for _ in executor.map(multi_agent_rag.process_query, queries):
            pass


if __name__ == "__main__":
    main()
